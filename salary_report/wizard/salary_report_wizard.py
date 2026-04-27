from odoo import api, models, fields
from odoo.exceptions import UserError
from io import BytesIO
import base64
from collections import defaultdict
import calendar


class SalaryReportWizard(models.TransientModel):
    _name = "salary.report.wizard"
    _description = "Salary Report"

    company_id = fields.Many2one(
        "res.company",
        string="Company",
        required=True,
        default=lambda self: self.env.company,
    )
    date_from = fields.Date(string="From Date", required=True)
    date_to = fields.Date(string="To Date", required=True)

    all_employee = fields.Boolean(string="All Employees")
    employee_ids = fields.Many2many(
        "hr.employee",
        string="Employees",
        domain="[('active', '=', True), ('company_id', '=', company_id)]",
    )

    @api.onchange("company_id")
    def _onchange_company_id(self):
        allowed_employees = self._get_employees_for_company()
        self.employee_ids = self.employee_ids.filtered(lambda emp: emp in allowed_employees)

    def _get_company_scoped_self(self):
        self.ensure_one()
        return self.with_company(self.company_id).with_context(
            allowed_company_ids=[self.company_id.id]
        )

    def _get_employees_for_company(self):
        scoped_self = self._get_company_scoped_self()
        return scoped_self.env["hr.employee"].search([
            ("active", "=", True),
            ("company_id", "=", scoped_self.company_id.id),
        ])

    def _get_selected_employees(self):
        self.ensure_one()
        employees = self._get_employees_for_company()
        if self.all_employee:
            return employees
        if not self.employee_ids:
            raise UserError("Please select employees or tick All Employees.")
        return employees.filtered(lambda emp: emp.id in self.employee_ids.ids)

    def _get_report_employee(self, employee):
        self.ensure_one()
        return employee.sudo().with_company(self.company_id).with_context(
            allowed_company_ids=[self.company_id.id]
        )

    def action_bulk_enable_pf(self):
        scoped_self = self._get_company_scoped_self()
        contracts = scoped_self.env['hr.contract'].search([
            ('state', '=', 'open'),
            ('l10n_in_provident_fund', '=', False),
            ('company_id', '=', scoped_self.company_id.id),
        ])

        if not contracts:
            raise UserError(
                "No contracts found to update. All contracts already have PF enabled or no active contracts exist."
            )

        contracts.write({'l10n_in_provident_fund': True})

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'PF Enabled Successfully',
                'message': f'Provident Fund enabled for {len(contracts)} active contract(s)',
                'type': 'success',
                'sticky': False,
            }
        }

    # ------------------------------------------------------------------
    # Salary component extraction helpers
    # Priority: exact code match → name keyword match → category match
    # This makes the wizard work across ALL salary structures without
    # any hardcoded assumptions about component values.
    # ------------------------------------------------------------------

    def _by_code(self, lines, *codes):
        upper = {c.upper() for c in codes}
        return sum(l.total for l in lines if l.code and l.code.upper() in upper)

    def _by_name(self, lines, *keywords):
        kws = [k.lower() for k in keywords]
        return sum(
            l.total for l in lines
            if l.name and any(k in l.name.lower() for k in kws)
        )

    def _by_category(self, lines, *cat_codes):
        upper = {c.upper() for c in cat_codes}
        return sum(
            l.total for l in lines
            if l.category_id and l.category_id.code and l.category_id.code.upper() in upper
        )

    def _get_basic(self, lines):
        v = self._by_code(lines, 'BASIC', 'BASIC_SALARY', 'BASIC_SAL')
        if not v:
            v = self._by_name(lines, 'basic salary')
        if not v:
            # category code BASIC
            v = self._by_category(lines, 'BASIC')
        return v

    def _get_hra(self, lines):
        v = self._by_code(lines, 'HRA', 'HOUSE_RENT', 'HOUSE_RENT_ALLOWANCE')
        if not v:
            v = self._by_name(lines, 'house rent allowance')
        return v

    def _get_conveyance(self, lines):
        v = self._by_code(lines, 'CONV', 'CONVEYANCE', 'CONVEYANCE_ALLOWANCE', 'CONVEYANCE_ALLOWNCE')
        if not v:
            v = self._by_name(lines, 'conveyance')
        return v

    def _get_lta(self, lines):
        v = self._by_code(lines, 'LTA', 'LEAVE_TRAVEL', 'LEAVE_TRAVEL_ALLOWANCE', 'LTA_REIMB')
        if not v:
            v = self._by_name(lines, 'leave travel allowance', 'lta reimb')
        return v

    def _get_bonus(self, lines):
        v = self._by_code(lines, 'BONUS', 'GROSS_BONUS')
        if not v:
            v = self._by_category(lines, 'BONUS')
        return v

    def _get_other_allowance(self, lines, basic, hra, conveyance, lta, bonus):
        """
        Other Allowance = everything in ALLOWANCE category
        that is NOT Basic / HRA / Conveyance / LTA / Bonus.

        Strategy:
        1. Try known codes first.
        2. Try name match.
        3. Fallback: sum ALL allowance-category lines minus the known ones.
           This works for ANY salary structure (India Regular Pay, German TMT, etc.)
        """
        # 1. Known codes
        v = self._by_code(
            lines,
            'OTHER_ALW', 'OTHER_ALLOWANCE', 'OTHER', 'OTH_ALW', 'OTHERALLOW',
            'STD_ALW', 'STANDARD_ALLOWANCE',          # India Regular Pay: Standard Allowance
            'SPECIAL_ALW', 'SUPPLEMENTARY_ALW',        # other common names
        )
        if v:
            return v

        # 2. Name match
        v = self._by_name(
            lines,
            'other allowance', 'standard allowance',
            'special allowance', 'supplementary allowance',
        )
        if v:
            return v

        # 3. Fallback: sum all ALW/ALLOWANCE category lines, exclude known components
        allowance_total = self._by_category(lines, 'ALW', 'ALLOWANCE', 'ALLOW')
        other = allowance_total - hra - conveyance - lta
        if other > 0:
            return other

        # 4. Last resort: gross minus known earning components
        gross_line = self._get_gross(lines)
        if gross_line:
            remainder = gross_line - basic - hra - conveyance - lta - bonus
            return max(remainder, 0)

        return 0

    def _get_gross(self, lines):
        # NOTE: Do NOT include TAXABLE_SALARY here — that is the PF-wage base,
        # not the Total Gross Earning shown on the payslip.
        v = self._by_code(
            lines,
            'GROSS', 'GROSS_EARN', 'TOTAL_GROSS', 'GROSS_SALARY',
            'TOTAL_GROSS_EARN', 'TOTAL_GROSS_EARNING', 'TOTAL_EARN',
            'GROSS_BONUS',   # some structures roll bonus into gross line
        )
        if not v:
            # Name match covers "Total Gross Earning" line in German TMT structure
            v = self._by_name(
                lines,
                'total gross earning', 'gross earning', 'gross salary', 'total gross'
            )
        return v

    def _get_net(self, lines):
        # NET_SALARY / NET_SAL first
        v = self._by_code(lines, 'NET_SALARY', 'NET_SAL')
        if v:
            return v
        # Name match 'net salary' — picks Net Salary line only, not In Hand Salary
        v = self._by_name(lines, 'net salary', 'net payable')
        if v:
            return v
        # NET code (safe — In Hand Salary uses INHAND code, not NET)
        v = self._by_code(lines, 'NET')
        if v:
            return v
        # Last resort: In Hand Salary
        v = self._by_code(lines, 'INHAND')
        if not v:
            v = self._by_name(lines, 'in hand salary')
        return v

    def _get_pf(self, lines):
        v = self._by_code(lines, 'PF', 'PROVIDENT_FUND', 'EPF', 'PF_EMP', 'EMPPF')
        if not v:
            v = self._by_name(lines, 'provident fund', 'employee pf')
        return abs(v)

    def _get_esic(self, lines):
        v = self._by_code(lines, 'ESIC', 'ESI', 'ESIC_EMP', 'ESI_EMP')
        if not v:
            v = self._by_name(lines, 'esic', 'esi')
        return abs(v)

    def _get_component_pt(self, lines):
        v = self._by_code(lines, 'PT', 'PROF_TAX', 'PROFESSIONAL_TAX', 'P_TAX', 'P_TAX_DED')
        if not v:
            v = self._by_name(lines, 'professional tax', 'prof tax')
        return abs(v)

    def _get_pf_salary(self, lines, basic, conveyance, lta):
        v = self._by_code(lines, 'PF_SALARY', 'PF_WAGE', 'TAXABLE_SALARY', 'PF_SAL')
        if not v:
            v = self._by_name(lines, 'pf salary', 'taxable salary')
        if not v and basic:
            v = min(basic + conveyance + lta, 15000)
        return v

    def action_generate_excel(self):
        self = self._get_company_scoped_self()
        if self.date_from > self.date_to:
            raise UserError("From Date cannot be greater than To Date")

        employees = self._get_selected_employees()

        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Salary report"

        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        blue_fill = PatternFill("solid", fgColor="5B9BD5")
        light_blue_fill = PatternFill("solid", fgColor="87CEFA")
        grey_fill = PatternFill("solid", fgColor="D9D9D9")

        header_font = Font(bold=True)

        headers = [
            "Emp Code", "Other Bank Ac No", "ICICI Ac No",
            "PF?", "ESIC?", "UAN", "ESIC NO", "Employee Name",
            "New Gross", "Department Name", "Designation Name",
            "CTC salary", "GROSS SALARY", "Basic", "HRA",
            "Conveyance", "LTA", "Other Allowance", "TOTAL",
            "Bonus", "Payable", "PF SALARY",
            "Paid Days", "Basic", "HRA", "Conveyance", "LTA",
            "Other Allowance", "Bonus", "AREARS", "Earning Sal",
            "PF SALARY", "PF", "ESIC", "PT", "L.W.F",
            "Advance Bank", "Advance cash", "TDS",
            "Total Deduction", "Net Salary", "Remark", "Status"
        ]

        total_cols = len(headers)


        # Get the year and month from date_from
        year = self.date_from.year
        month = self.date_from.month
        # Calculate the number of days in the selected month
        days_in_month = calendar.monthrange(year, month)[1]


        company = self.company_id
        partner = company.partner_id
        address = ", ".join(filter(None, [
            partner.street or "", partner.street2 or "",
            partner.city or "",
            partner.state_id.name if partner.state_id else "",
            partner.zip or "",
        ]))

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        ws["A1"] = f"{company.name}\n{address}"
        ws["A1"].font = Font(size=20, bold=True)
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws["A1"].fill = light_blue_fill
        ws.row_dimensions[1].height = 80
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 41

        ws.merge_cells("A2:B2")
        ws["A2"] = "Month:"
        ws["A2"].fill = grey_fill
        ws["C2"] = self.date_from.strftime("%B %Y")

        ws.merge_cells("E2:F2")
        ws["E2"] = "Days:"
        ws["E2"].fill = grey_fill
        ws["G2"] = (self.date_to - self.date_from).days + 1

        column_widths = {
            "A": 12, "B": 20, "C": 20, "D": 17, "E": 17,
            "F": 17, "G": 17, "H": 26, "I": 11, "J": 13,
            "K": 13, "L": 11, "M": 12, "N": 10, "O": 10,
            "P": 12, "Q": 10, "R": 11, "S": 11, "T": 10,
            "U": 10, "V": 10, "W": 10, "X": 10, "Y": 10,
            "Z": 10, "AA": 10, "AB": 10, "AC": 10, "AD": 10,
            "AE": 10, "AF": 10, "AG": 10, "AH": 10, "AI": 10,
            "AJ": 10, "AK": 10, "AL": 10, "AM": 10,
            "AN": 10, "AO": 10, "AP": 14, "AQ": 12,
        }

        header_row = 3
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(header_row, col, header)
            cell.font = header_font
            cell.alignment = align_center
            cell.fill = blue_fill
            cell.border = border
            ws.column_dimensions[get_column_letter(col)].width = column_widths.get(get_column_letter(col), 18)

        payslips = self.env["hr.payslip"].search([
            ("employee_id", "in", employees.ids),
            ("date_from", "<=", self.date_to),
            ("date_to", ">=", self.date_from),
            ("state", "in", ["done", "paid"]),
        ])

        slip_map = defaultdict(list)
        for slip in payslips:
            slip_map[slip.employee_id.id].append(slip)

        row = header_row + 1

        for emp in employees:
            emp_report = self._get_report_employee(emp)
            emp_slips = slip_map.get(emp.id, [])

            # Paid days
            paid_days = sum(
                wd.number_of_days
                for slip in emp_slips
                for wd in slip.worked_days_line_ids
                if wd.is_paid
            )

            payslip_lines = self.env['hr.payslip.line']
            for slip in emp_slips:
                payslip_lines |= slip.line_ids


            basic       = self._get_basic(payslip_lines)
            hra         = self._get_hra(payslip_lines)
            conveyance  = self._get_conveyance(payslip_lines)
            lta         = self._get_lta(payslip_lines)
            bonus       = self._get_bonus(payslip_lines)
            other_allowance = self._get_other_allowance(
                payslip_lines, basic, hra, conveyance, lta, bonus
            )

            gross = self._get_gross(payslip_lines)
            if not gross:
                gross = basic + hra + conveyance + lta + other_allowance + bonus

            net        = self._get_net(payslip_lines)
            pf         = self._get_pf(payslip_lines)
            esic       = self._get_esic(payslip_lines)
            pf_salary  = self._get_pf_salary(payslip_lines, basic, conveyance, lta)

            pf_flag   = "Yes" if pf   != 0 else "No"
            esic_flag = "Yes" if esic != 0 else "No"

            # Bank routing
            other_bank_ac = icici_ac = ""
            if emp_report.bank_account_id and emp_report.bank_account_id.bank_id:
                bank_name = (emp_report.bank_account_id.bank_id.name or "").upper()
                acc_no = emp_report.bank_account_id.acc_number or ""
                if "ICICI" in bank_name or "ICIC" in bank_name:
                    icici_ac = acc_no
                else:
                    other_bank_ac = acc_no


            earnings_total = gross if gross else (basic + hra + conveyance + lta + other_allowance)

            data = [
                emp_report.employee_code or "",
                other_bank_ac,
                icici_ac,
                pf_flag,
                esic_flag,
                emp_report.l10n_in_uan or "Not Available",
                emp_report.l10n_in_esic_number or "Not Available",
                emp_report.name,
                gross,
                emp_report.department_id.name if emp_report.department_id else "",
                emp_report.job_id.name if emp_report.job_id else "",
                emp_report.contract_id.final_yearly_costs if emp_report.contract_id else 0,
                gross,
                basic,
                hra,
                conveyance,
                lta,
                other_allowance,
                earnings_total,
                bonus,
                net,
                pf_salary,
            ]

            final_row_data = data + [""] * (total_cols - len(data))

            for col, value in enumerate(final_row_data, start=1):
                ws.cell(row, col, value).border = border

            ws[f"W{row}"]  = paid_days
            ws[f"X{row}"]  = f"=ROUND(N{row}*W{row}/{days_in_month},0)"
            ws[f"Y{row}"]  = f"=ROUND(O{row}*W{row}/{days_in_month},0)"
            ws[f"Z{row}"]  = f"=ROUND(P{row}*W{row}/{days_in_month},0)"
            ws[f"AA{row}"] = f"=ROUND(Q{row}*W{row}/{days_in_month},0)"
            ws[f"AB{row}"] = f"=ROUND(R{row}*W{row}/{days_in_month},0)"
            ws[f"AC{row}"] = f"=ROUND(T{row}*W{row}/{days_in_month},0)"
            ws[f"AD{row}"] = ""
            company_name = (self.company_id.name or "").upper()
            if "GREEN" in company_name or "GGSPL" in company_name:
                ws[f"AE{row}"] = f"=ROUND(S{row}*W{row}/{days_in_month},0)"
            else:
                ws[f"AE{row}"] = f"=SUM(X{row}:AD{row})"

            ws[f"AF{row}"] = pf_salary
            ws[f"AG{row}"] = pf
            ws[f"AH{row}"] = esic
            ws[f"AI{row}"] = abs(self._get_component_pt(payslip_lines))
            ws[f"AJ{row}"] = ""
            ws[f"AK{row}"] = f"=IFERROR(VLOOKUP(A{row},ADVANCE!$A$4:$E$309,5,0),0)"
            ws[f"AL{row}"] = ""
            ws[f"AM{row}"] = ""
            ws[f"AN{row}"] = f"=AG{row}+AH{row}+AI{row}+AJ{row}+AK{row}+AL{row}+AM{row}"
            ws[f"AO{row}"] = f"=U{row}"
            ws[f"AP{row}"] = ""
            ws[f"AQ{row}"] = ""

            row += 1

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        attachment = self.env["ir.attachment"].create({
            "name": f"Salary_Report_{self.date_from.strftime('%B %Y')}.xlsx",
            "type": "binary",
            "datas": base64.b64encode(output.read()),
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "company_id": self.company_id.id,
            "res_model": self._name,
            "res_id": self.id,
        })

        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}?download=true",
            "target": "self",
        }
