# -*- coding: utf-8 -*-
from odoo import api, models, fields
from odoo.exceptions import UserError
from io import BytesIO
import base64
from collections import defaultdict
import calendar
from datetime import datetime, time, timedelta
from pytz import timezone, UTC


class SalaryReport2(models.Model):
    _name = "salary.report.2"
    _description = "Monthly Salary Report Snapshot"
    _order = "date_from desc, id desc"

    name = fields.Char(string="Report Name", required=True, compute="_compute_name", store=True, readonly=False)
    company_id = fields.Many2one(
        "res.company",
        string="Company",
        required=True,
        default=lambda self: self.env.company,
        readonly=True,
        states={'draft': [('readonly', False)]},
        index=True,
    )
    date_from = fields.Date(string="From Date", required=True)
    date_to = fields.Date(string="To Date", required=True)
    state = fields.Selection(
        [
            ("draft", "Draft"),
            ("calculated", "Calculated"),
            ("locked", "Locked / Closed"),
        ],
        string="Status",
        default="draft",
        required=True,
        tracking=True,
    )
    all_employee = fields.Boolean(string="All Employees", default=True)
    employee_ids = fields.Many2many(
        "hr.employee",
        string="Specific Employees",
        domain="[('company_id', '=', company_id)]",
        context={'active_test': False},
    )

    line_ids = fields.One2many(
        "salary.report.line.2",
        "report_id",
        string="Salary Sheet Lines",
        copy=True,
    )

    total_employees = fields.Integer(string="Total Employees", compute="_compute_totals", store=True)
    total_gross_salary = fields.Float(string="Total Gross Salary", compute="_compute_totals", store=True)
    total_net_pay = fields.Float(string="Total Net Pay", compute="_compute_totals", store=True)

    @api.depends("date_from", "date_to", "company_id")
    def _compute_name(self):
        for rec in self:
            if rec.date_from:
                month_name = rec.date_from.strftime("%B %Y")
                rec.name = f"Salary Sheet - {month_name}"
            else:
                rec.name = "Salary Sheet"

    @api.depends("line_ids", "line_ids.gross_salary", "line_ids.net_pay")
    def _compute_totals(self):
        for rec in self:
            rec.total_employees = len(rec.line_ids)
            rec.total_gross_salary = sum(rec.line_ids.mapped("gross_salary"))
            rec.total_net_pay = sum(rec.line_ids.mapped("net_pay"))

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        today = fields.Date.today()
        from odoo.tools import date_utils
        res["date_from"] = date_utils.start_of(today, "month")
        res["date_to"] = date_utils.end_of(today, "month")
        return res

    def action_lock(self):
        for rec in self:
            rec.state = "locked"

    def action_reset_draft(self):
        for rec in self:
            rec.state = "draft"

    # ------------------------------------------------------------------
    # Employee filtering logic:
    # 1. Exclude employees joined AFTER date_to.
    # 2. Include resigned/inactive employees if they worked during period.
    # ------------------------------------------------------------------
    def _get_eligible_employees(self):
        self.ensure_one()
        all_emps = self.env["hr.employee"].with_context(active_test=False).search([
            ("company_id", "=", self.company_id.id),
        ])
        
        eligible = self.env["hr.employee"]
        for emp in all_emps:
            # Check joining date: if joined after date_to, exclude
            if emp.joining_date and emp.joining_date > self.date_to:
                continue

            # Check departure date: if departed before date_from, exclude
            if emp.departure_date and emp.departure_date < self.date_from:
                continue

            eligible |= emp

        if not self.all_employee:
            if not self.employee_ids:
                raise UserError("Please select specific employees or check 'All Employees'.")
            eligible = eligible.filtered(lambda e: e.id in self.employee_ids.ids)

        return eligible

    # ------------------------------------------------------------------
    # Helpers for extraction
    # ------------------------------------------------------------------
    def _by_code(self, lines, *codes):
        upper = {c.upper() for c in codes}
        return sum(l.total for l in lines if l.code and l.code.upper() in upper)

    def _by_name(self, lines, *keywords):
        kws = [k.lower() for k in keywords]
        return sum(
            l.total for l in lines
            if l.name and any(k in l.name.lower() for k in kws)
        )

    def _by_category(self, lines, *cat_codes):
        upper = {c.upper() for c in cat_codes}
        return sum(
            l.total for l in lines
            if l.category_id and l.category_id.code and l.category_id.code.upper() in upper
        )

    def _get_from_contract(self, employee, *codes_or_names):
        contract = employee.contract_id
        if not contract or not contract.salary_structure_line_ids:
            return 0.0
        upper_terms = {t.upper() for t in codes_or_names}
        lines = contract.salary_structure_line_ids
        matching = lines.filtered(lambda l: l.code and l.code.upper() in upper_terms)
        if matching:
            return sum(matching.mapped("amount_monthly"))
        matching = lines.filtered(lambda l: l.name and l.name.upper() in upper_terms)
        if matching:
            return sum(matching.mapped("amount_monthly"))
        matching = lines.filtered(lambda l: l.name and any(term in l.name.upper() for term in upper_terms))
        if matching:
            return sum(matching.mapped("amount_monthly"))
        return 0.0

    def _get_basic(self, lines):
        v = self._by_code(lines, "BASIC", "BASIC_SALARY", "BASIC_SAL")
        if not v:
            v = self._by_name(lines, "basic salary")
        if not v:
            v = self._by_category(lines, "BASIC")
        return v

    def _get_hra(self, lines):
        v = self._by_code(lines, "HRA", "HOUSE_RENT", "HOUSE_RENT_ALLOWANCE")
        if not v:
            v = self._by_name(lines, "house rent allowance")
        return v

    def _get_bonus(self, lines, employee=None):
        v = self._by_code(lines, "BONUS", "BOUNS", "GROSS_BONUS")
        if not v:
            v = self._by_name(lines, "bonus", "bouns")
        if not v:
            v = self._by_category(lines, "BONUS")
        v = abs(v)
        if not v and employee:
            v = self._get_from_contract(employee, "BONUS", "BOUNS")
        return v

    def _get_pf(self, lines, employee=None):
        v = self._by_code(lines, "PF", "PROVIDENT_FUND", "EPF", "PF_EMP", "EMPPF")
        if not v:
            v = self._by_name(lines, "provident fund", "employee pf")
        v = abs(v)
        if not v and employee:
            v = self._get_from_contract(employee, "PF", "PROVIDENT FUND", "EPF")
        return v

    def _get_component_pt(self, lines, employee=None):
        v = self._by_code(lines, "PT", "PROF_TAX", "PROFESSIONAL_TAX", "P_TAX", "P_TAX_DED")
        if not v:
            v = self._by_name(lines, "professional tax", "prof tax")
        v = abs(v)
        if not v and employee:
            v = self._get_from_contract(employee, "PT", "PROFESSIONAL TAX")
        return v

    def _get_lwf(self, lines, employee=None):
        v = self._by_code(lines, "LWF", "LABOUR_WELFARE", "LWF_EMP")
        if not v:
            v = self._by_name(lines, "lwf", "labour welfare")
        v = abs(v)
        if not v and employee:
            v = self._get_from_contract(employee, "LWF", "LABOUR WELFARE")
        return v

    def _get_it(self, lines, employee=None):
        v = self._by_code(lines, "TDS", "IT", "INCOME_TAX", "TAX", "TDS_DED")
        if not v:
            v = self._by_name(lines, "tds", "income tax", "tax")
        v = abs(v)
        if not v and employee:
            v = self._get_from_contract(employee, "TDS", "IT", "INCOME TAX", "TAX")
        return v

    def _get_staff_loan(self, lines, employee=None):
        v = self._by_code(lines, "LOAN", "STAFF_LOAN", "LOAN_DED")
        if not v:
            v = self._by_name(lines, "loan", "staff loan")
        v = abs(v)
        if not v and employee:
            v = self._get_from_contract(employee, "LOAN", "STAFF LOAN")
        return v

    def _get_cash_advance(self, lines, employee=None):
        v = self._by_code(lines, "ADVANCE", "CASH_ADV", "ADV_DED")
        if not v:
            v = self._by_name(lines, "advance", "cash advance")
        v = abs(v)
        if not v and employee:
            v = self._get_from_contract(employee, "ADVANCE", "CASH ADVANCE")
        return v

    def _get_other_deduction(self, lines, employee=None):
        v = self._by_code(lines, "OTHER", "OTHER_DED", "OTHER_DEDUCTION")
        if not v:
            v = self._by_name(lines, "other deduction")
        v = abs(v)
        if not v and employee:
            v = self._get_from_contract(employee, "OTHER", "OTHER DEDUCTION")
        return v

    def _get_safety_equip(self, lines, employee=None):
        v = self._by_code(lines, "SAFETY", "UNIFORM", "SAFETY_EQUIP", "SAFETY_DED", "UNIFORM_DED")
        if not v:
            v = self._by_name(lines, "safety", "uniform", "safety equipment")
        v = abs(v)
        if not v and employee:
            v = self._get_from_contract(employee, "SAFETY", "UNIFORM", "SAFETY EQUIPMENT")
        return v

    def _get_salary_paid(self, lines, employee=None):
        v = self._by_code(lines, "SALARY_PAID", "PREV_PAID", "PAID_DED")
        if not v:
            v = self._by_name(lines, "salary paid", "previously paid")
        v = abs(v)
        if not v and employee:
            v = self._get_from_contract(employee, "SALARY PAID", "PREVIOUSLY PAID")
        return v

    def _get_incentive(self, lines, employee=None):
        v = self._by_code(lines, "INCENTIVE", "INC", "BONUS_INC")
        if not v:
            v = self._by_name(lines, "incentive")
        v = abs(v)
        if not v and employee:
            v = self._get_from_contract(employee, "INCENTIVE", "INC")
        return v

    def _get_leave_type_code(self, leave_type_name):
        name_lower = (leave_type_name or '').lower()
        if 'casual' in name_lower or name_lower == 'cl':
            return 'CL'
        elif 'earned' in name_lower or 'annual' in name_lower or name_lower == 'el':
            return 'EL'
        elif 'sick' in name_lower or name_lower == 'sl':
            return 'SL'
        elif 'unpaid' in name_lower or 'lwp' in name_lower or 'loss of pay' in name_lower:
            return 'UL'
        elif 'management' in name_lower or 'mgmt' in name_lower:
            return 'ML'
        elif 'on duty' in name_lower or name_lower == 'od' or 'onduty' in name_lower:
            return 'OD'
        elif 'half' in name_lower:
            return 'HD'
        else:
            return leave_type_name[:3].upper() if leave_type_name else 'L'

    def _get_daily_attendance_status(self, employee, start_date, end_date):
        calendar_obj = employee.resource_calendar_id or employee.company_id.resource_calendar_id
        if not calendar_obj:
            return {}

        tz = timezone(calendar_obj.tz) if calendar_obj.tz else UTC
        start = tz.localize(datetime.combine(start_date, time.min))
        stop = tz.localize(datetime.combine(end_date, time.max))

        attendances = self.env['hr.attendance'].sudo().search([
            ('employee_id', '=', employee.id),
            ('check_in', '>=', start.astimezone(UTC).replace(tzinfo=None)),
            ('check_in', '<=', stop.astimezone(UTC).replace(tzinfo=None)),
        ])

        attendances_by_day = defaultdict(list)
        for att in attendances:
            day = att.check_in.astimezone(tz).date() if att.check_in else None
            if day:
                attendances_by_day[day].append(att)

        leaves = self.env['hr.leave'].sudo().search([
            ('employee_id', '=', employee.id),
            ('state', '=', 'validate'),
            ('date_from', '<', stop.astimezone(UTC).replace(tzinfo=None)),
            ('date_to', '>', start.astimezone(UTC).replace(tzinfo=None)),
        ])

        leave_info_by_day = {}
        for leave in leaves:
            leave_start = leave.date_from.astimezone(tz).date()
            leave_end = leave.date_to.astimezone(tz).date()
            leave_type_name = leave.holiday_status_id.name if leave.holiday_status_id else 'Leave'
            leave_type_code = self._get_leave_type_code(leave_type_name)
            is_half_day = leave.request_unit_half if hasattr(leave, 'request_unit_half') else False
            
            current = leave_start
            while current <= leave_end:
                if start_date <= current <= end_date:
                    leave_info_by_day[current] = {
                        'type_code': leave_type_code,
                        'is_half_day': is_half_day,
                    }
                current += timedelta(days=1)

        holidays = self.env['resource.calendar.leaves'].sudo().search([
            ('resource_id', '=', False),
            ('calendar_id', '=', calendar_obj.id),
            ('date_from', '<', stop.astimezone(UTC).replace(tzinfo=None)),
            ('date_to', '>', start.astimezone(UTC).replace(tzinfo=None)),
        ])

        holiday_days = set()
        for holiday in holidays:
            holiday_start = holiday.date_from.astimezone(tz).date()
            holiday_stop = holiday.date_to.astimezone(tz).date()
            current = max(holiday_start, start_date)
            end_date_eff = min(holiday_stop, end_date)
            while current <= end_date_eff:
                holiday_days.add(current)
                current += timedelta(days=1)

        daily_status = {}
        date_range = [start_date + timedelta(days=x) for x in range((end_date - start_date).days + 1)]
        today = fields.Date.today()

        for day in date_range:
            is_weekly_off = False
            if hasattr(employee, '_get_weekly_off'):
                is_weekly_off = bool(employee.sudo()._get_weekly_off(day))

            if day in holiday_days:
                status = 'H'
            elif day in leave_info_by_day:
                leave_data = leave_info_by_day[day]
                if leave_data['is_half_day']:
                    if day in attendances_by_day:
                        status = 'P'
                    else:
                        status = 'L'
                else:
                    status = leave_data['type_code']
            elif is_weekly_off:
                status = 'W'
            else:
                day_start = tz.localize(datetime.combine(day, time.min))
                day_end = tz.localize(datetime.combine(day, time.max))
                try:
                    attendance_intervals = calendar_obj._attendance_intervals_batch(
                        day_start, day_end, resources=employee.resource_id
                    )[employee.resource_id.id]
                except:
                    attendance_intervals = []

                if not attendance_intervals:
                    status = 'W'
                else:
                    if day in attendances_by_day:
                        status = 'P'
                    elif day > today:
                        status = ''
                    else:
                        status = 'A'

            daily_status[day] = status

        return daily_status

    def _get_daily_overtime(self, employee, start_date, end_date):
        calendar_obj = employee.resource_calendar_id or employee.company_id.resource_calendar_id
        if not calendar_obj:
            return {}

        tz = timezone(calendar_obj.tz) if calendar_obj.tz else UTC
        start = tz.localize(datetime.combine(start_date, time.min))
        stop = tz.localize(datetime.combine(end_date, time.max))

        work_entries = self.env['hr.work.entry'].sudo().search([
            ('employee_id', '=', employee.id),
            ('date_start', '<=', stop.astimezone(UTC).replace(tzinfo=None)),
            ('date_stop', '>=', start.astimezone(UTC).replace(tzinfo=None)),
            ('state', '!=', 'cancelled'),
        ])

        ot_type = self.env.ref('hr_work_entry.overtime_work_entry_type', raise_if_not_found=False)
        weekly_off_type = employee.company_id.weekly_off_work_entry_type_id if hasattr(employee.company_id, 'weekly_off_work_entry_type_id') else False
        
        ot_types = self.env['hr.work.entry.type']
        if ot_type:
            ot_types |= ot_type
        if weekly_off_type:
            ot_types |= weekly_off_type
        overtime_code_types = self.env['hr.work.entry.type'].sudo().search([('code', '=', 'OVERTIME')])
        if overtime_code_types:
            ot_types |= overtime_code_types

        ot_entries = work_entries.filtered(lambda we: we.work_entry_type_id in ot_types)

        daily_ot = defaultdict(float)
        for we in ot_entries:
            we_day = we.date_start.astimezone(tz).date()
            if start_date <= we_day <= end_date:
                duration = we.duration or ((we.date_stop - we.date_start).total_seconds() / 3600.0)
                daily_ot[we_day] += duration

        return daily_ot

    # ------------------------------------------------------------------
    # Action: Fetch and Calculate Data into persistent line snapshot
    # ------------------------------------------------------------------
    def action_fetch_and_compute(self):
        for report in self:
            if report.state == "locked":
                raise UserError("Cannot recalculate a locked report.")

            if report.date_from > report.date_to:
                raise UserError("From Date cannot be greater than To Date.")

            # Clear existing lines
            report.line_ids.unlink()

            employees = report._get_eligible_employees()
            days_in_month = calendar.monthrange(report.date_from.year, report.date_from.month)[1]

            payslips = self.env["hr.payslip"].search([
                ("employee_id", "in", employees.ids),
                ("date_from", "<=", report.date_to),
                ("date_to", ">=", report.date_from),
                ("state", "in", ["done", "paid", "verify"]),
            ])
            slip_map = defaultdict(list)
            for slip in payslips:
                slip_map[slip.employee_id.id].append(slip)

            line_vals = []
            sr_no = 1
            for emp in employees:
                emp_slips = slip_map.get(emp.id, [])
                daily_status = report._get_daily_attendance_status(emp, report.date_from, report.date_to)
                daily_ot = report._get_daily_overtime(emp, report.date_from, report.date_to)

                payslip_lines = self.env['hr.payslip.line']
                for slip in emp_slips:
                    payslip_lines |= slip.line_ids

                # Bank account details
                sbi_ac = sbi_ifsc = hdfc_ac = hdfc_ifsc = ""
                if emp.bank_account_id and emp.bank_account_id.bank_id:
                    bank_name = (emp.bank_account_id.bank_id.name or '').upper()
                    acc_num = emp.bank_account_id.acc_number or ''
                    ifsc_val = emp.bank_account_id.bank_id.bic or ''
                    if 'SBI' in bank_name or 'STATE BANK' in bank_name:
                        sbi_ac = acc_num
                        sbi_ifsc = ifsc_val
                    elif 'HDFC' in bank_name:
                        hdfc_ac = acc_num
                        hdfc_ifsc = ifsc_val

                # Attendance days dict
                att_dict = {}
                for day_num in range(1, 32):
                    field_name = f"day_{day_num}"
                    if day_num <= days_in_month:
                        current_date = report.date_from.replace(day=day_num)
                        att_dict[field_name] = daily_status.get(current_date, '')
                    else:
                        att_dict[field_name] = ''

                # OT days dict
                ot_dict = {}
                tot_ot = 0.0
                for day_num in range(1, 32):
                    field_name = f"ot_day_{day_num}"
                    if day_num <= days_in_month:
                        current_date = report.date_from.replace(day=day_num)
                        ot_hrs = daily_ot.get(current_date, 0.0)
                        ot_dict[field_name] = ot_hrs
                        tot_ot += ot_hrs
                    else:
                        ot_dict[field_name] = 0.0

                # Salary specs
                ctc = emp.contract_id.wage or 0.0 if emp.contract_id else 0.0
                basic_spec = ctc * 0.60
                hra_spec = ctc * 0.40

                # Pay days
                if emp_slips:
                    pay_days = sum(
                        wd.number_of_days
                        for slip in emp_slips
                        for wd in slip.worked_days_line_ids
                        if wd.is_paid
                    )
                else:
                    pay_days = sum(
                        1 for d_val, status in daily_status.items()
                        if status in ['P', 'W', 'H'] or (status and status not in ['A', 'UL'])
                    )

                # Payroll components
                calc_basic = round((basic_spec * pay_days / days_in_month), 0) if days_in_month else 0.0
                calc_hra = round((hra_spec * pay_days / days_in_month), 0) if days_in_month else 0.0
                tot_salary = calc_basic + calc_hra
                incentive = report._get_incentive(payslip_lines, emp)
                gross_salary = tot_salary + incentive

                pt = report._get_component_pt(payslip_lines, emp)
                pf = report._get_pf(payslip_lines, emp)
                lwf = report._get_lwf(payslip_lines, emp)
                bonus = report._get_bonus(payslip_lines, emp)
                it = report._get_it(payslip_lines, emp)
                staff_loan = report._get_staff_loan(payslip_lines, emp)
                cash_advance = report._get_cash_advance(payslip_lines, emp)
                other_ded = report._get_other_deduction(payslip_lines, emp)
                safety_equip = report._get_safety_equip(payslip_lines, emp)
                salary_paid = report._get_salary_paid(payslip_lines, emp)

                tot_deductions = pt + pf + lwf + it + staff_loan + cash_advance + other_ded + safety_equip + salary_paid
                net_pay = gross_salary - tot_deductions

                vals = {
                    "report_id": report.id,
                    "sr_no": sr_no,
                    "employee_id": emp.id,
                    "employee_code": emp.employee_code or "",
                    "employee_name": emp.name or "",
                    "father_name": emp.father_name or "" if hasattr(emp, "father_name") else "",
                    "sbi_ac": sbi_ac,
                    "sbi_ifsc": sbi_ifsc,
                    "hdfc_ac": hdfc_ac,
                    "hdfc_ifsc": hdfc_ifsc,
                    "uan_no": emp.l10n_in_uan or "" if hasattr(emp, "l10n_in_uan") else "",
                    "mobile_no": emp.mobile_phone or "",
                    "dob": emp.birthday,
                    "doj": emp.joining_date if hasattr(emp, "joining_date") else False,
                    "department_name": emp.department_id.name or "",
                    "designation_name": emp.job_id.name or "",
                    "ctc": ctc,
                    "basic_spec": basic_spec,
                    "hra_spec": hra_spec,
                    "month_days": days_in_month,
                    "ot_hours": tot_ot,
                    "pay_days": pay_days,
                    "basic": calc_basic,
                    "hra": calc_hra,
                    "total_salary": tot_salary,
                    "incentive": incentive,
                    "gross_salary": gross_salary,
                    "pt": pt,
                    "pf": pf,
                    "lwf": lwf,
                    "bonus": bonus,
                    "it": it,
                    "staff_loan": staff_loan,
                    "cash_advance": cash_advance,
                    "other_deduction": other_ded,
                    "safety_equipment": safety_equip,
                    "salary_paid": salary_paid,
                    "net_pay": net_pay,
                }
                vals.update(att_dict)
                vals.update(ot_dict)
                line_vals.append(vals)
                sr_no += 1

            self.env["salary.report.line.2"].create(line_vals)
            report.state = "calculated"

    # ------------------------------------------------------------------
    # Action: Generate XLSX directly from saved UI database lines
    # ------------------------------------------------------------------
    def action_generate_xlsx(self):
        self.ensure_one()
        if not self.line_ids:
            raise UserError("No data found to generate XLSX. Please click 'Fetch & Calculate Data' first.")

        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Complete Salary Sheet"

        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )
        peach_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        blue_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        # Row 1: Title
        ws.merge_cells("A1:CU1")
        title_cell = ws["A1"]
        title_cell.value = f"COMPLETE SALARY FOR THE MONTH OF {self.date_from.strftime('%B-%Y').upper()}"
        title_cell.font = Font(name="Calibri", size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = peach_fill
        ws.row_dimensions[1].height = 40

        # Row 2: Headers
        headers = [
            "NO", "EMP CODE", "NAME", "FATHER/HUSBAND NAME", "SBI AC", "IFSC CODE",
            "HDFCAC", "HDFC IFSC", "UANNO", "MOBILE NO", "DOB", "DOJ",
            "DEPARTMENT", "DESIGNATION"
        ]
        headers += [str(day) for day in range(1, 32)]  # Attendance 1 to 31
        headers += ["SALARY/ CTC", "BASIC", "HRA"]
        headers += [str(day) for day in range(1, 32)]  # OT 1 to 31
        headers += [
            "MONTH OF DAY", "OT HOURS", "PAY DAYS", "BASIC", "HRA",
            "TOTAL SALARY", "INCENTIVE", "GROSS SALARY", "PT", "PF",
            "LWF", "BONUS", "IT", "STAFF LOAN", "CASH ADVANCE", "OTHER",
            "SAFETY EQUIP ENT & UNIFOR", "SALARY PAID", "NET PAY", "SIGNATURE"
        ]

        header_font = Font(name="Calibri", size=10, bold=True)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 35

        for col_idx, h_text in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=h_text)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            if 15 <= col_idx <= 45:
                cell.fill = blue_fill
            elif 49 <= col_idx <= 79:
                cell.fill = green_fill
            else:
                cell.fill = peach_fill

        # Data Rows from stored database lines
        p_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        p_font = Font(name='Calibri', size=10, color='006100', bold=True)
        w_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        w_font = Font(name='Calibri', size=10, color='9C5700', bold=True)
        a_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        a_font = Font(name='Calibri', size=10, color='9C0006', bold=True)
        h_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
        h_font = Font(name='Calibri', size=10, color='1F4E79', bold=True)
        l_fill = PatternFill(start_color='E4DFEC', end_color='E4DFEC', fill_type='solid')
        l_font = Font(name='Calibri', size=10, color='7030A0', bold=True)

        row = 3
        for line in self.line_ids.sorted(key=lambda l: l.sr_no):
            ws.cell(row, 1, line.sr_no).alignment = Alignment(horizontal='center')
            ws.cell(row, 2, line.employee_code or '').alignment = Alignment(horizontal='center')
            ws.cell(row, 3, line.employee_name or '').alignment = Alignment(horizontal='left')
            ws.cell(row, 4, line.father_name or '').alignment = Alignment(horizontal='left')
            ws.cell(row, 5, line.sbi_ac or '').alignment = Alignment(horizontal='center')
            ws.cell(row, 6, line.sbi_ifsc or '').alignment = Alignment(horizontal='center')
            ws.cell(row, 7, line.hdfc_ac or '').alignment = Alignment(horizontal='center')
            ws.cell(row, 8, line.hdfc_ifsc or '').alignment = Alignment(horizontal='center')
            ws.cell(row, 9, line.uan_no or '').alignment = Alignment(horizontal='center')
            ws.cell(row, 10, line.mobile_no or '').alignment = Alignment(horizontal='center')

            dob_str = line.dob.strftime('%d-%m-%Y') if line.dob else ''
            doj_str = line.doj.strftime('%d-%m-%Y') if line.doj else ''
            ws.cell(row, 11, dob_str).alignment = Alignment(horizontal='center')
            ws.cell(row, 12, doj_str).alignment = Alignment(horizontal='center')
            ws.cell(row, 13, line.department_name or '').alignment = Alignment(horizontal='left')
            ws.cell(row, 14, line.designation_name or '').alignment = Alignment(horizontal='left')

            for col_idx in range(1, 15):
                cell = ws.cell(row, col_idx)
                cell.font = Font(name='Calibri', size=10)
                cell.border = thin_border

            # Attendance Days (15 to 45)
            for day_num in range(1, 32):
                col_idx = 14 + day_num
                cell = ws.cell(row, col_idx)
                cell.border = thin_border
                cell.font = Font(name='Calibri', size=10)
                cell.alignment = Alignment(horizontal='center')
                status = getattr(line, f"day_{day_num}", '') or ''
                cell.value = status

                if status == 'P':
                    cell.fill = p_fill
                    cell.font = p_font
                elif status == 'W':
                    cell.fill = w_fill
                    cell.font = w_font
                elif status == 'A':
                    cell.fill = a_fill
                    cell.font = a_font
                elif status == 'H':
                    cell.fill = h_fill
                    cell.font = h_font
                elif status:
                    cell.fill = l_fill
                    cell.font = l_font

            # CTC & specs (46 to 48)
            ws.cell(row, 46, line.ctc).alignment = Alignment(horizontal='right')
            ws.cell(row, 47, line.basic_spec).alignment = Alignment(horizontal='right')
            ws.cell(row, 48, line.hra_spec).alignment = Alignment(horizontal='right')
            for col_idx in range(46, 49):
                cell = ws.cell(row, col_idx)
                cell.font = Font(name='Calibri', size=10)
                cell.border = thin_border
                cell.number_format = '#,##0.00'

            # Overtime days (49 to 79)
            for day_num in range(1, 32):
                col_idx = 48 + day_num
                cell = ws.cell(row, col_idx)
                cell.border = thin_border
                cell.font = Font(name='Calibri', size=10)
                cell.alignment = Alignment(horizontal='center')
                ot_val = getattr(line, f"ot_day_{day_num}", 0.0) or 0.0
                cell.value = ot_val if ot_val > 0 else ''

            # Payroll & Net Pay (80 to 99)
            ws.cell(row, 80, line.month_days).alignment = Alignment(horizontal='center')
            ws.cell(row, 81, f"=SUM(AW{row}:CA{row})").alignment = Alignment(horizontal='center')
            ws.cell(row, 82, line.pay_days).alignment = Alignment(horizontal='center')
            ws.cell(row, 83, line.basic).alignment = Alignment(horizontal='right')
            ws.cell(row, 84, line.hra).alignment = Alignment(horizontal='right')
            ws.cell(row, 85, line.total_salary).alignment = Alignment(horizontal='right')
            ws.cell(row, 86, line.incentive).alignment = Alignment(horizontal='right')
            ws.cell(row, 87, line.gross_salary).alignment = Alignment(horizontal='right')
            ws.cell(row, 88, line.pt).alignment = Alignment(horizontal='right')
            ws.cell(row, 89, line.pf).alignment = Alignment(horizontal='right')
            ws.cell(row, 90, line.lwf).alignment = Alignment(horizontal='right')
            ws.cell(row, 91, line.bonus).alignment = Alignment(horizontal='right')
            ws.cell(row, 92, line.it).alignment = Alignment(horizontal='right')
            ws.cell(row, 93, line.staff_loan).alignment = Alignment(horizontal='right')
            ws.cell(row, 94, line.cash_advance).alignment = Alignment(horizontal='right')
            ws.cell(row, 95, line.other_deduction).alignment = Alignment(horizontal='right')
            ws.cell(row, 96, line.safety_equipment).alignment = Alignment(horizontal='right')
            ws.cell(row, 97, line.salary_paid).alignment = Alignment(horizontal='right')
            ws.cell(row, 98, line.net_pay).alignment = Alignment(horizontal='right')
            ws.cell(row, 99, line.signature or "").alignment = Alignment(horizontal='center')

            for col_idx in range(80, 100):
                cell = ws.cell(row, col_idx)
                cell.font = Font(name='Calibri', size=10)
                cell.border = thin_border
                if col_idx in [83, 84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 94, 95, 96, 97, 98]:
                    cell.number_format = '#,##0.00'

            row += 1

        # Totals Row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
        total_cell = ws.cell(row, 1, "TOTAL")
        total_cell.font = Font(name='Calibri', size=10, bold=True)
        total_cell.alignment = Alignment(horizontal='center')
        total_cell.fill = peach_fill

        for col_idx in range(1, 100):
            cell = ws.cell(row, col_idx)
            cell.border = thin_border
            cell.font = Font(name='Calibri', size=10, bold=True)
            if col_idx < 15:
                cell.fill = peach_fill

        sum_cols = [46, 47, 48, 81, 82, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 94, 95, 96, 97, 98]
        for col_idx in sum_cols:
            col_letter = get_column_letter(col_idx)
            cell = ws.cell(row, col_idx, f"=SUM({col_letter}3:{col_letter}{row-1})")
            if col_idx in [46, 47, 48, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 94, 95, 96, 97, 98]:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='center')

        # Column Widths
        column_widths = {
            "A": 5, "B": 12, "C": 25, "D": 25, "E": 18, "F": 15,
            "G": 18, "H": 15, "I": 18, "J": 15, "K": 12, "L": 12,
            "M": 18, "N": 18, "AT": 15, "AU": 12, "AV": 12,
            "CB": 12, "CC": 12, "CD": 12, "CE": 12, "CF": 12,
            "CG": 15, "CH": 12, "CI": 15, "CJ": 10, "CK": 12,
            "CL": 10, "CM": 12, "CN": 10, "CO": 12, "CP": 12,
            "CQ": 10, "CR": 22, "CS": 12, "CT": 15, "CU": 15
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        for day_num in range(1, 32):
            ws.column_dimensions[get_column_letter(14 + day_num)].width = 4
            ws.column_dimensions[get_column_letter(48 + day_num)].width = 4

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        attachment = self.env["ir.attachment"].create({
            "name": f"{self.name.replace(' ', '_')}.xlsx",
            "type": "binary",
            "datas": base64.b64encode(output.read()),
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "company_id": self.company_id.id,
            "res_model": self._name,
            "res_id": self.id,
        })

        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}?download=true",
            "target": "self",
        }


class SalaryReportLine2(models.Model):
    _name = "salary.report.line.2"
    _description = "Salary Report Line Snapshot"
    _order = "sr_no, id"

    report_id = fields.Many2one(
        "salary.report.2", string="Salary Report", required=True, ondelete="cascade"
    )
    company_id = fields.Many2one(
        "res.company",
        string="Company",
        related="report_id.company_id",
        store=True,
        readonly=True,
        index=True,
    )
    sr_no = fields.Integer(string="No")
    employee_id = fields.Many2one("hr.employee", string="Employee", context={'active_test': False})
    employee_code = fields.Char(string="Emp Code")
    employee_name = fields.Char(string="Name")
    father_name = fields.Char(string="Father/Husband Name")
    sbi_ac = fields.Char(string="SBI A/C")
    sbi_ifsc = fields.Char(string="SBI IFSC")
    hdfc_ac = fields.Char(string="HDFC A/C")
    hdfc_ifsc = fields.Char(string="HDFC IFSC")
    uan_no = fields.Char(string="UAN No")
    mobile_no = fields.Char(string="Mobile No")
    dob = fields.Date(string="DOB")
    doj = fields.Date(string="DOJ")
    department_name = fields.Char(string="Department")
    designation_name = fields.Char(string="Designation")

    # Attendance Snapshot Days 1 to 31
    day_1 = fields.Char(string="1")
    day_2 = fields.Char(string="2")
    day_3 = fields.Char(string="3")
    day_4 = fields.Char(string="4")
    day_5 = fields.Char(string="5")
    day_6 = fields.Char(string="6")
    day_7 = fields.Char(string="7")
    day_8 = fields.Char(string="8")
    day_9 = fields.Char(string="9")
    day_10 = fields.Char(string="10")
    day_11 = fields.Char(string="11")
    day_12 = fields.Char(string="12")
    day_13 = fields.Char(string="13")
    day_14 = fields.Char(string="14")
    day_15 = fields.Char(string="15")
    day_16 = fields.Char(string="16")
    day_17 = fields.Char(string="17")
    day_18 = fields.Char(string="18")
    day_19 = fields.Char(string="19")
    day_20 = fields.Char(string="20")
    day_21 = fields.Char(string="21")
    day_22 = fields.Char(string="22")
    day_23 = fields.Char(string="23")
    day_24 = fields.Char(string="24")
    day_25 = fields.Char(string="25")
    day_26 = fields.Char(string="26")
    day_27 = fields.Char(string="27")
    day_28 = fields.Char(string="28")
    day_29 = fields.Char(string="29")
    day_30 = fields.Char(string="30")
    day_31 = fields.Char(string="31")

    # Base CTC Specs
    ctc = fields.Float(string="Salary / CTC")
    basic_spec = fields.Float(string="Basic Spec")
    hra_spec = fields.Float(string="HRA Spec")

    # Overtime Snapshot Days 1 to 31
    ot_day_1 = fields.Float(string="OT 1")
    ot_day_2 = fields.Float(string="OT 2")
    ot_day_3 = fields.Float(string="OT 3")
    ot_day_4 = fields.Float(string="OT 4")
    ot_day_5 = fields.Float(string="OT 5")
    ot_day_6 = fields.Float(string="OT 6")
    ot_day_7 = fields.Float(string="OT 7")
    ot_day_8 = fields.Float(string="OT 8")
    ot_day_9 = fields.Float(string="OT 9")
    ot_day_10 = fields.Float(string="OT 10")
    ot_day_11 = fields.Float(string="OT 11")
    ot_day_12 = fields.Float(string="OT 12")
    ot_day_13 = fields.Float(string="OT 13")
    ot_day_14 = fields.Float(string="OT 14")
    ot_day_15 = fields.Float(string="OT 15")
    ot_day_16 = fields.Float(string="OT 16")
    ot_day_17 = fields.Float(string="OT 17")
    ot_day_18 = fields.Float(string="OT 18")
    ot_day_19 = fields.Float(string="OT 19")
    ot_day_20 = fields.Float(string="OT 20")
    ot_day_21 = fields.Float(string="OT 21")
    ot_day_22 = fields.Float(string="OT 22")
    ot_day_23 = fields.Float(string="OT 23")
    ot_day_24 = fields.Float(string="OT 24")
    ot_day_25 = fields.Float(string="OT 25")
    ot_day_26 = fields.Float(string="OT 26")
    ot_day_27 = fields.Float(string="OT 27")
    ot_day_28 = fields.Float(string="OT 28")
    ot_day_29 = fields.Float(string="OT 29")
    ot_day_30 = fields.Float(string="OT 30")
    ot_day_31 = fields.Float(string="OT 31")

    # Summary Totals & Payroll Components
    month_days = fields.Integer(string="Month Days")
    ot_hours = fields.Float(string="OT Hours")
    pay_days = fields.Float(string="Pay Days")
    basic = fields.Float(string="Basic")
    hra = fields.Float(string="HRA")
    total_salary = fields.Float(string="Total Salary")
    incentive = fields.Float(string="Incentive")
    gross_salary = fields.Float(string="Gross Salary")
    pt = fields.Float(string="PT")
    pf = fields.Float(string="PF")
    lwf = fields.Float(string="LWF")
    bonus = fields.Float(string="Bonus")
    it = fields.Float(string="IT")
    staff_loan = fields.Float(string="Staff Loan")
    cash_advance = fields.Float(string="Cash Advance")
    other_deduction = fields.Float(string="Other Deduction")
    safety_equipment = fields.Float(string="Safety Equip & Uniform")
    salary_paid = fields.Float(string="Salary Paid")
    net_pay = fields.Float(string="Net Pay")
    signature = fields.Char(string="Signature")

    # Onchanges for dynamic UI live editing
    @api.onchange(
        'day_1', 'day_2', 'day_3', 'day_4', 'day_5', 'day_6',
        'day_7', 'day_8', 'day_9', 'day_10', 'day_11', 'day_12',
        'day_13', 'day_14', 'day_15', 'day_16', 'day_17', 'day_18',
        'day_19', 'day_20', 'day_21', 'day_22', 'day_23', 'day_24',
        'day_25', 'day_26', 'day_27', 'day_28', 'day_29', 'day_30', 'day_31'
    )
    def _onchange_attendance_days(self):
        for line in self:
            days = [
                line.day_1, line.day_2, line.day_3, line.day_4, line.day_5, line.day_6,
                line.day_7, line.day_8, line.day_9, line.day_10, line.day_11, line.day_12,
                line.day_13, line.day_14, line.day_15, line.day_16, line.day_17, line.day_18,
                line.day_19, line.day_20, line.day_21, line.day_22, line.day_23, line.day_24,
                line.day_25, line.day_26, line.day_27, line.day_28, line.day_29, line.day_30, line.day_31
            ]
            count = 0.0
            for d in days:
                if not d:
                    continue
                d_str = str(d).strip().upper()
                if d_str in ['P', 'W', 'H', 'CL', 'EL', 'SL', 'ML', 'OD']:
                    count += 1.0
                elif d_str == 'HD':
                    count += 0.5
                elif d_str not in ['A', 'UL']:
                    count += 1.0
            line.pay_days = count
            line._recompute_salary_from_pay_days()

    @api.onchange('pay_days', 'ctc', 'month_days')
    def _onchange_pay_days(self):
        for line in self:
            line._recompute_salary_from_pay_days()

    def _recompute_salary_from_pay_days(self):
        for line in self:
            ctc = line.ctc or 0.0
            line.basic_spec = ctc * 0.60
            line.hra_spec = ctc * 0.40
            m_days = line.month_days or 30
            if m_days > 0:
                line.basic = round((line.basic_spec * (line.pay_days or 0.0) / m_days), 0)
                line.hra = round((line.hra_spec * (line.pay_days or 0.0) / m_days), 0)
            else:
                line.basic = 0.0
                line.hra = 0.0
            line.total_salary = (line.basic or 0.0) + (line.hra or 0.0)
            line.gross_salary = line.total_salary + (line.incentive or 0.0)
            line._recompute_net_pay()

    @api.onchange(
        'ot_day_1', 'ot_day_2', 'ot_day_3', 'ot_day_4', 'ot_day_5', 'ot_day_6',
        'ot_day_7', 'ot_day_8', 'ot_day_9', 'ot_day_10', 'ot_day_11', 'ot_day_12',
        'ot_day_13', 'ot_day_14', 'ot_day_15', 'ot_day_16', 'ot_day_17', 'ot_day_18',
        'ot_day_19', 'ot_day_20', 'ot_day_21', 'ot_day_22', 'ot_day_23', 'ot_day_24',
        'ot_day_25', 'ot_day_26', 'ot_day_27', 'ot_day_28', 'ot_day_29', 'ot_day_30', 'ot_day_31'
    )
    def _onchange_ot_days(self):
        for line in self:
            line.ot_hours = sum([
                line.ot_day_1 or 0.0, line.ot_day_2 or 0.0, line.ot_day_3 or 0.0, line.ot_day_4 or 0.0,
                line.ot_day_5 or 0.0, line.ot_day_6 or 0.0, line.ot_day_7 or 0.0, line.ot_day_8 or 0.0,
                line.ot_day_9 or 0.0, line.ot_day_10 or 0.0, line.ot_day_11 or 0.0, line.ot_day_12 or 0.0,
                line.ot_day_13 or 0.0, line.ot_day_14 or 0.0, line.ot_day_15 or 0.0, line.ot_day_16 or 0.0,
                line.ot_day_17 or 0.0, line.ot_day_18 or 0.0, line.ot_day_19 or 0.0, line.ot_day_20 or 0.0,
                line.ot_day_21 or 0.0, line.ot_day_22 or 0.0, line.ot_day_23 or 0.0, line.ot_day_24 or 0.0,
                line.ot_day_25 or 0.0, line.ot_day_26 or 0.0, line.ot_day_27 or 0.0, line.ot_day_28 or 0.0,
                line.ot_day_29 or 0.0, line.ot_day_30 or 0.0, line.ot_day_31 or 0.0,
            ])

    @api.onchange('basic', 'hra')
    def _onchange_salary(self):
        for line in self:
            line.total_salary = (line.basic or 0.0) + (line.hra or 0.0)
            line.gross_salary = line.total_salary + (line.incentive or 0.0)
            line._recompute_net_pay()

    @api.onchange('total_salary', 'incentive')
    def _onchange_gross(self):
        for line in self:
            line.gross_salary = (line.total_salary or 0.0) + (line.incentive or 0.0)
            line._recompute_net_pay()

    @api.onchange(
        'gross_salary', 'pt', 'pf', 'lwf', 'bonus', 'it',
        'staff_loan', 'cash_advance', 'other_deduction', 'safety_equipment', 'salary_paid'
    )
    def _onchange_deductions(self):
        for line in self:
            line._recompute_net_pay()

    def _recompute_net_pay(self):
        for line in self:
            line.net_pay = (line.gross_salary or 0.0) - (
                (line.pt or 0.0) + (line.pf or 0.0) + (line.lwf or 0.0) +
                (line.it or 0.0) + (line.staff_loan or 0.0) + (line.cash_advance or 0.0) +
                (line.other_deduction or 0.0) + (line.safety_equipment or 0.0) + (line.salary_paid or 0.0)
            )

